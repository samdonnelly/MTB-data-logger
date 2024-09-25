#================================================================================
# File: mtbdl_data_analysis.py 
# 
# Author: Sam Donnelly (samueldonnelly11@gmail.com)
# 
# Description: MTBDL data analysis script 
#
# Date: July 27, 2023 
#================================================================================

#================================================================================
# Notes 

# Notes: 
# - A 'file_locations.py' script must be created in the 'analysis' folder (not 
#   a sub-folder within) for each local copy of this repository. This script 
#   contains the file locations of data used and generated by the analysis. It 
#   is not tracked by git as directories will change between machines. For file 
#   location names, see the comments and code below. 

# UI 
# - Choose mode --> format log files, analyze data or coordinate plot 
# - Format log files; 
#   - Select a date - check for folder matching date 
#   - Select a log file range - check that those files exist (could also be 1 file) 
#   - Format all valid files sequentially: 
#     - Need to format GPS coordinates, IMU readings (correct for resting state), 
#       calculate wheel speed, create time data (?), suspension travel calculation 
#       (ADC voltage to position), UTC time to local time. 
#     - Suspension parameters recorded and pasted per log. 
#     - ALl logs from a folder saved to one spreadsheet. 
#     - One data log data set per one spreadsheet sheet. Sheet tab names are the 
# - Analyze data 
#   - Data must be formatted first. 
#   - Trailmarker points plotted on every plot. 
#   - Plot Latitude and longitude on separate plots against time. 
# - Coordinate plot 
#   - Formatting will have to be done first. 
#   - Take GPS coordinates and plot on a map. 

#================================================================================

#================================================================================
# Includes 

# Files and functions 
from file_locations import mtbdl_logs_raw 

# Libraries 
from openpyxl import Workbook, load_workbook 
import atexit 
import os 

import matplotlib.animation as animation 

#================================================================================


#================================================================================
# Global variables 

# String data 
mode_index = ("1", "2", "3") 
modes = ("Format", "Analysis", "Coordinate plot") 
mode_str = "\r\n" + mode_index[0] + ". " + modes[0] + "\r\n" + \
                    mode_index[1] + ". " + modes[1] + "\r\n" + \
                    mode_index[2] + ". " + modes[2] + "\r\n" 

# User prompts 
mode_prompt = "Mode: "
date_prompt = "\nDate (YYYY-MM-DD): "
low_log_prompt = "Lower log index: " 
high_log_prompt = "Upper log index: " 
terminate_prompt = "exit"

#================================================================================


#================================================================================
# Functions 

##
# brief: Code called upon termination of the script 
# 
# description: Code that should be run before the script terminates should be put here. 
#              This code will only be run once the program has begun the termination 
#              process. 
##
def exit_handler(): 
    print("\nProgram terminated.\n") 


##
# brief: Checks user input for a valid entry 
# 
# description: The function prints a prompt (specified in the arguments) to the 
#              console and takes the user input. This input is first checked 
#              against the 'check' argument, which if matched, will then 
#              terminate the script. If there is no match then the 'data_type' 
#              argument is used to determine how to interpret the user input. 
#              If the user input aligns with the data types specified then 
#              'True' is returned along with the user input formatted 
#              according to the data type. If the input doesn't match the 
#              data types then 'False' is returned and the user input returned 
#              is irrelevant.  
# 
# @param prompt : string printed to console to prompt the user as needed 
# @param check : string to check user input against for terminating the program 
# @param data_type : expected data type input from the user 
# @return 1 : True if user input matches expected format, False otherwise 
# @return 2 : contents of the user input formatted as needed if it's a valid input 
##
def user_input(
    prompt, 
    check, 
    data_type): 

    # Get the user input 
    command = input(prompt) 
        
    # Check for exit 
    # String comparision 
    for i in range(len(check)): 
        try: 
            if (command[i].lower() != check[i]):
                break 
        except IndexError: 
            break 
    
    if (i == (len(check)-1)): 
        exit() 
    
    # Try converting the number and check for errors 
    if (data_type is float): 
        # Try converting to float and check for errors 
        try: 
            value = float(command) 
            return True, value
        except ValueError: 
            print("\nFloat conversion failed.\n") 
    
    elif (data_type is int): 
        # Try converting to int and check for errors 
        try: 
            value = int(float(command)) 
            return True, value
        except ValueError: 
            print("\nInteger conversion failed.\n") 
            
    else: 
        # Interpreted as a string 
        return True, command
            
    return False, 0


##
# brief: Select the data operation to perform 
# 
# description: Allows the user to choose between formatting raw data log files, 
#              running the analysis on a set of formatted data logs or plotting the 
#              GPS coodinates on a map. This is the default user prompt for the 
#              script whenno other action is being used. 
##
def mode_select(): 
    while (True): 
        print(mode_str) 
        valid, user_str = user_input(mode_prompt, terminate_prompt, str) 
        if (valid): 
            # Interpret the input 
            user_str = user_str.lower()
            if (user_str == mode_index[0] or user_str == modes[0].lower()): 
                format_data() 
            elif (user_str == mode_index[1] or user_str == modes[1].lower()): 
                analyze_data() 
            elif (user_str == mode_index[2] or user_str == modes[2].lower()): 
                coord_plot() 


##
# brief: Format data log data 
# 
# description: Prompts the user to specify a set of raw data log files, checks if the 
#              files exist, then formats the data into an Excel sheet so it can be 
#              used for data analysis. If any of the specified data log files don't 
#              exist then the request will be aborted. If the Excel file does not 
#              already exist then it will be created first. Only one sheet per log 
#              file number will be created so logs that have already been formatted 
#              will be overwritten if requested again. Data is intended to be 
#              formatted in the most useful or human readable way possible. For 
#              example, IMU data is not formatted into xyz compoents but rather 
#              bike pitch and roll as well as total acceleration. 
##
def format_data(): 
    # Get the log data from the user 
    while (True): 
        valid, user_in = user_input(date_prompt, terminate_prompt, str) 
        if (valid): 
            log_folder = mtbdl_logs_raw + "/" + user_in 
            if (os.path.isdir(log_folder) is not True): 
                print("No logs for that date.") 
                return 
            log_date = user_in
            break 
    
    # Get the log index range from the user 
    while (True): 
        valid, user_in = user_input(low_log_prompt, terminate_prompt, int) 
        if (valid): 
            log_low_index = user_in 
            break 
    while (True): 
        valid, user_in = user_input(high_log_prompt, terminate_prompt, int) 
        if (valid): 
            log_high_index = user_in 
            break 

    # Check that all the logs exist 
    for i in range(log_low_index, log_high_index+1): 
        log_file_raw = log_folder + "/" + log_date + "-log" + str(i) + ".txt"
        if (os.path.exists(log_file_raw) is not True): 
            print("Some log files in the specified range don't exist.") 
            return 

    # All files exist 

    # Check for the existance of an Excel file 
    log_file = log_folder + "/" + log_date + "-log-data.xlsx" 
    if (os.path.exists(log_file) is not True): 
        # If it does not exist then create it and name the first sheet 
        wb = Workbook()
        ws = wb.active 
        ws.title = "log-" + str(log_low_index) 
    else: 
        # If it exists then open it 
        wb = load_workbook(log_file) 
        ws = wb.active 
    
    # Check for the existance of log sheets - create them if they don't exist 
    sheet_names = wb.sheetnames 
    for i in range(log_low_index, log_high_index+1): 
        sheet_check = 0 
        sheet = "log-" + str(i) 

        for name in sheet_names: 
            if (name == sheet): 
                sheet_check += 1 
                break 
        
        if (not sheet_check): 
            # sheet does not exist - create it 
            wb.create_sheet(sheet) 

    # Write data to each log sheet 
    for i in range(log_low_index, log_high_index+1): 
        # Set the log file and sheet name to access 
        log_file_raw = log_folder + "/" + log_date + "-log" + str(i) + ".txt" 
        sheet = "log-" + str(i) 

        # Open the requested log file and sheet 
        file_obj = open(log_file_raw, "r") 
        ws = wb[sheet] 

        # Data tracking 
        time = 0.0 
        time_step = 0.0 
        trailmarker = [[], []] 
        fork = [] 
        shock = [] 
        wheel_speed = [] 
        imu = [] 
        gps = [] 

        # Read, set and write calculation/formatting data 
        line_data = file_obj.readline() 
        while(line_data != "Data log:"): 
            # If fork parameters 
            # If shock parameters 
            # If IMU offsets 
            # If potentiometer offsets 
            # If UTC time 
            # If logging params 
            time_step = 0.0 
            # Read next line 
            line_data = file_obj.readline() 

        # Read, format and write log data 
        line_data = file_obj.readline() 
        while(line_data != "End"): 
            # Check trail marker 
            # Update fork and shock arrays 
            # Check for wheel speed - if available then update time + data 
            # Check for IMU - if available then update time + data 
            # Check for GPS - if available then update time + data 
            # Update universal time tracker and read next line 
            time += time_step 
            line_data = file_obj.readline() 

        ws["B" + str(i)] = i 
    
    # Save the log file to a specified name 
    wb.save(log_file) 


##
# brief: Perform analysis on formatted data logs 
# 
# description: Takes formatted data log data and present it to the user in a readable
#              and visual format so they can understand their riding style and bike 
#              setup. 
##
def analyze_data(): 
    print("\nAnalyze") 
    return 


##
# brief: Plot recorded GPS coordinates on a map 
# 
# description: 
##
def coord_plot(): 
    print("\nPlot") 
    return 

#================================================================================


#================================================================================
# Run 

#==================================================
# Setup 

# Configure the exit handler 
atexit.register(exit_handler) 

# Exit statement 
print("\n-----------------------------------------------------")
print("Type 'exit' into any prompt to terminate the program.")

#==================================================

#==================================================
# Main loop 

while (True): 
    mode_select() 

#==================================================

#================================================================================


#================================================================================
# https://matplotlib.org/stable/users/explain/animations/animations.html 

# fig, ax = plt.subplots()
# line = ax.plot(fork_travel_data[0], pot_voltage_data[0])[0] 
# ax.set_xlabel("Fork Travel (mm)") 
# ax.set_ylabel("Fork Pot Voltage (V)") 
# ax.set_title("Fork Travel <--> Potentiometer Voltage")
# ax.set(xlim=[0, 200], ylim=[1.2, 3.0])

# def update(frame):
#     line.set_xdata(fork_travel_data[:frame])
#     line.set_ydata(pot_voltage_data[:frame])
#     return line 

# ani = animation.FuncAnimation(fig=fig, func=update, frames=400, interval=1)
# plt.show()

#================================================================================
